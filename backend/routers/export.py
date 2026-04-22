from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from datetime import datetime

from backend.database import get_db
from backend.models.exam import Exam
from backend.models.result import Result
from backend.models.student import Student

router = APIRouter()

@router.get("/excel/{exam_id}")
async def export_excel(exam_id: int, db: Session = Depends(get_db)):
    exam = db.query(Exam).filter(Exam.id == exam_id).first()
    if not exam:
        raise HTTPException(status_code=404, detail="Examen no encontrado")
        
    results = db.query(Result).filter(Result.exam_id == exam_id).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resultados"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="CC0000", end_color="CC0000", fill_type="solid")
    center_align = Alignment(horizontal="center")
    
    # Encabezados
    headers = ["Puesto", "Nombre del Estudiante", "DNI", "Correctas", "Incorrectas", "Blancas", "Nota Final"]
    for col, text in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
    
    # Datos
    sorted_results = sorted(results, key=lambda x: x.score, reverse=True)
    for i, r in enumerate(sorted_results, 1):
        student = db.query(Student).filter(Student.id == r.student_id).first()
        ws.cell(row=i+1, column=1, value=i)
        ws.cell(row=i+1, column=2, value=student.name if student else "Desconocido")
        ws.cell(row=i+1, column=3, value=student.dni if student else "-")
        ws.cell(row=i+1, column=4, value=r.correct)
        ws.cell(row=i+1, column=5, value=r.wrong)
        ws.cell(row=i+1, column=6, value=r.blank)
        ws.cell(row=i+1, column=7, value=r.score)
    
    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = max_length + 2

    file_path = f"data/exports/resultados_{exam_id}.xlsx"
    wb.save(file_path)
    
    return FileResponse(file_path, filename=f"resultados_{exam.name}.xlsx")

@router.get("/pdf-boletas/{exam_id}")
async def export_pdf_boletas(exam_id: int, db: Session = Depends(get_db)):
    exam = db.query(Exam).filter(Exam.id == exam_id).first()
    if not exam:
        raise HTTPException(status_code=404, detail="Examen no encontrado")
        
    results = db.query(Result).filter(Result.exam_id == exam_id).all()
    file_path = f"data/exports/boletas_{exam_id}.pdf"
    
    c = canvas.Canvas(file_path, pagesize=A4)
    width, height = A4
    
    for r in results:
        student = db.query(Student).filter(Student.id == r.student_id).first()
        
        # Dibujar Boleta
        c.setFont("Helvetica-Bold", 16)
        c.setFillColorRGB(0.8, 0, 0)
        c.drawCentredString(width/2, height - 20*mm, "COLEGIO ALBERT EINSTEIN")
        
        c.setFont("Helvetica-Bold", 12)
        c.setFillColorRGB(0, 0, 0)
        c.drawCentredString(width/2, height - 30*mm, "BOLETA DE RESULTADOS OMR")
        
        c.setLineWidth(0.5)
        c.line(20*mm, height - 35*mm, width - 20*mm, height - 35*mm)
        
        # Información
        c.setFont("Helvetica", 10)
        c.drawString(25*mm, height - 45*mm, f"EXAMEN: {exam.name}")
        c.drawString(25*mm, height - 50*mm, f"MATERIA: {exam.subject}")
        c.drawString(25*mm, height - 55*mm, f"FECHA: {datetime.now().strftime('%d/%m/%Y')}")
        
        c.setFont("Helvetica-Bold", 11)
        c.drawString(25*mm, height - 65*mm, f"ESTUDIANTE: {student.name if student else 'Desconocido'}")
        c.drawString(25*mm, height - 70*mm, f"DNI: {student.dni if student else '-'}")
        
        # Cuadro de puntaje
        c.rect(25*mm, height - 100*mm, 160*mm, 20*mm)
        c.setFont("Helvetica-Bold", 14)
        c.drawString(30*mm, height - 92*mm, "PUNTAJE TOTAL:")
        c.drawRightString(180*mm, height - 92*mm, f"{r.score:.2f}")
        
        c.setFont("Helvetica", 10)
        c.drawString(30*mm, height - 110*mm, f"Respuestas Correctas: {r.correct}")
        c.drawString(30*mm, height - 115*mm, f"Respuestas Incorrectas: {r.wrong}")
        c.drawString(30*mm, height - 120*mm, f"Respuestas en Blanco: {r.blank}")
        
        c.showPage() # Nueva página por estudiante
        
    c.save()
    return FileResponse(file_path, filename=f"boletas_{exam.name}.pdf")
